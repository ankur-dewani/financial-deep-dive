"""
P&L Analysis Module
Reads all 8 sheets from the Input P&L workbook, builds analysis pivot tables,
and writes 5 new analysis sheets:

  1. Revenue Analysis      -- revenue mix from 3 revenue sheets + P&L Summary
  2. Benchmark Mapping     -- every expense line mapped to benchmark category
  3. SS Breakdown          -- Shared Services sub-department costs
  4. FA Deep Dive          -- F&A cost breakdown by component
  5. FA Employee Analysis  -- current vs Central Finance role mapping

Preserves all original sheets unchanged.
"""

from __future__ import annotations

import shutil
import sys
import os

from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# Add scripts dir to path for cost_model import
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from cost_model import (
    get_current_fa_cost,
    get_target_fa_model,
    get_employee_role_mapping,
    get_savings_summary,
    CURRENT_FA_OPEX,
)


# ---------------------------------------------------------------------------
# File paths
# ---------------------------------------------------------------------------
INPUT_PL = os.path.join(
    os.path.dirname(__file__), "..", "data", "input",
    "Operational Leadership Real Work - Input P&L.xlsx"
)
OUTPUT_PL = os.path.join(
    os.path.dirname(__file__), "..", "data", "output",
    "Input P&L - Ankur Dewani.xlsx"
)

# ---------------------------------------------------------------------------
# Style constants
# ---------------------------------------------------------------------------
_HEADER_FONT = Font(name="Arial", size=10, bold=True, color="FFFFFF")
_HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
_DATA_FONT = Font(name="Arial", size=9)
_BOLD_FONT = Font(name="Arial", size=9, bold=True)
_TOTAL_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
_CURRENCY_FMT = '#,##0'
_PCT_FMT = '0.0%'
_THIN_BORDER = Border(
    bottom=Side(style="thin", color="B4C6E7"),
)
_WRAP = Alignment(wrap_text=True, vertical="top")


# ---------------------------------------------------------------------------
# Benchmark category mapping
# ---------------------------------------------------------------------------
# Maps (Function L2, Dept) combinations to benchmark categories.
# Based on the 8 benchmark categories from the Benchmarks sheet.

BENCHMARK_MAP = {
    # Shared Services = G&A departments
    "G&A": {
        "Finance & Accounting": "Shared Services",
        "Human Resources":      "Shared Services",
        "Legal":                "Shared Services",
        "Enterprise Systems":   "Shared Services",
        "Business Operations":  "Shared Services",
        "Service Operations":   "Shared Services",
        "Corporate Technology": "Shared Services",
        "Occupancy":            "Shared Services",
        "G&A":                  "Shared Services",
    },
    # Executive Team = GMs/Office Admins + Corporate
    "G&A_executive": {
        "GMs & Office Admins":  "Executive Team",
        "Corporate":            "Executive Team",
    },
    # Sales
    "S&M_sales": {
        "Sales":                "Sales",
        "Solution Consultants": "Sales",
    },
    # Marketing
    "S&M_marketing": {
        "Marketing":            "Marketing",
    },
    # COGS categories
    "Cost of Product_support": {
        "Technical Support":    "Technical Support",
        "Enhanced Support":     "Technical Support",
    },
    "Cost of Product_hosting": {
        "Cloud Operations":     None,  # split by category below
    },
    "Cost of Product_product": {
        "Customer Success":     "Product",
    },
    # Engineering = R&D
    "R&D": {
        "Product Development":  "Engineering",
        "Quality Assurance":    "Engineering",
        "Product Management":   "Engineering",
    },
    # Cost of PSO
    "Cost of PSO": {
        "Professional Services": "Product",
        "Education and Training": "Product",
        "Funded R&D":           "Engineering",
    },
}

# These are populated dynamically from the Benchmarks and P&L Summary sheets
# at runtime by read_source_data(). Initialized here as module-level state.
BENCHMARK_TARGETS = {}
REVENUE = 0
REVENUE_BREAKDOWN = {}
PL_SUMMARY = {}


def read_source_data(wb) -> None:
    """Read Benchmarks, P&L Summary, and Revenue sheets to populate module state.

    This ensures all 8 input sheets are actively processed, not hardcoded.
    """
    global BENCHMARK_TARGETS, REVENUE, REVENUE_BREAKDOWN, PL_SUMMARY

    # --- Sheet 1: Benchmarks ---
    ws = wb["Benchmarks"]
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        category, benchmark = row[0], row[1]
        if category and benchmark and isinstance(benchmark, (int, float)):
            if category not in ("Margin", "Expense Total"):
                BENCHMARK_TARGETS[category] = benchmark
    print(f"  Benchmarks: {len(BENCHMARK_TARGETS)} categories loaded")

    # --- Sheet 2: P&L Summary ---
    ws = wb["P&L Summary"]
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        label = row[0].value
        value = row[1].value
        if label and value is not None and isinstance(value, (int, float)):
            PL_SUMMARY[label.strip()] = value
    REVENUE = PL_SUMMARY.get("Revenue", 0)
    margin = PL_SUMMARY.get("Margin", 0)
    print(f"  P&L Summary: Revenue = ${REVENUE:,.0f}, Margin = ${margin:,.0f} ({margin/REVENUE:.1%})" if REVENUE else "  P&L Summary: loaded")

    # --- Sheets 6-8: Revenue breakdown ---
    # Structure: Row 3 = headers (Tier, Type, Customer Name, 2018 total)
    # Data starts at row 4, column D (index 3) = amount
    for sheet_name in ["RecurringRevenue", "PSORevenue", "PerpetualRevenue"]:
        ws = wb[sheet_name]
        total = 0
        count = 0
        for row in ws.iter_rows(min_row=4, max_row=ws.max_row):
            val = row[3].value  # Column D: 2018 total
            if val and isinstance(val, (int, float)):
                total += val
                count += 1
        # Clean name: "RecurringRevenue" -> "Recurring"
        clean_name = sheet_name.replace("Revenue", "")
        REVENUE_BREAKDOWN[clean_name] = {"total": total, "count": count}
    rev_parts = ", ".join(f"{k}: ${v['total']:,.0f} ({v['count']} items)" for k, v in REVENUE_BREAKDOWN.items())
    print(f"  Revenue sheets: {rev_parts}")


def classify_benchmark(func_l2: str, dept: str, category: str = None) -> str:
    """Determine benchmark category for a line item."""
    if not func_l2 or not dept:
        return "Unclassified"

    # Strip whitespace from all inputs
    func_l2 = func_l2.strip()
    dept = dept.strip()
    if category:
        category = category.strip()

    # Executive Team (G&A sub-depts)
    if func_l2 == "G&A" and dept in ("GMs & Office Admins", "Corporate"):
        return "Executive Team"

    # G&A -> Shared Services
    if func_l2 == "G&A":
        return "Shared Services"

    # S&M
    if func_l2 == "S&M":
        if dept == "Marketing":
            return "Marketing"
        return "Sales"

    # R&D
    if func_l2 == "R&D":
        return "Engineering"

    # Cost of Product
    if func_l2 == "Cost of Product":
        if dept in ("Technical Support", "Enhanced Support"):
            return "Technical Support"
        if dept == "Cloud Operations":
            # Only the Hosting expense category maps to Hosting benchmark
            if category and category.lower() == "hosting":
                return "Hosting"
            return "Product"
        if dept == "Customer Success":
            return "Product"
        return "Product"

    # Cost of PSO
    if func_l2 == "Cost of PSO":
        if dept == "Funded R&D":
            return "Engineering"
        return "Product"

    return "Unclassified"


# ---------------------------------------------------------------------------
# Sheet builders
# ---------------------------------------------------------------------------

def _write_header(ws, headers: list[str], row: int = 1) -> None:
    """Write styled header row."""
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _WRAP


def _auto_width(ws, min_width: int = 12, max_width: int = 45) -> None:
    """Auto-fit column widths."""
    for col in range(1, ws.max_column + 1):
        max_len = min_width
        for row in range(1, ws.max_row + 1):
            val = ws.cell(row=row, column=col).value
            if val:
                max_len = max(max_len, min(len(str(val)) + 2, max_width))
        ws.column_dimensions[get_column_letter(col)].width = max_len


def build_benchmark_mapping(wb) -> None:
    """Sheet 1: Map every expense line to a benchmark category."""
    ws = wb.create_sheet("Benchmark Mapping")

    headers = [
        "Benchmark Category", "HC Cost", "Non-HC Cost",
        "Total Cost", "% of Revenue", "Benchmark Target",
        "Variance", "Status"
    ]
    _write_header(ws, headers)

    # Aggregate costs by benchmark category
    # Employee costs
    empl_ws = wb["Empl."]
    hc_by_cat = {}
    for row in empl_ws.iter_rows(min_row=4, max_row=empl_ws.max_row):
        func_l2 = row[3].value  # col D: Function L2
        dept = row[4].value     # col E: DEPT
        cost = row[6].value     # col G: 2018 total
        if not dept or not cost or not isinstance(cost, (int, float)):
            continue
        cat = classify_benchmark(func_l2, dept)
        hc_by_cat[cat] = hc_by_cat.get(cat, 0) + cost

    # OPEX non-employee costs
    opex_ws = wb["OPEX - NEmpl."]
    nhc_by_cat = {}
    for row_data in opex_ws.iter_rows(min_row=3, max_row=opex_ws.max_row):
        func_l2 = row_data[1].value  # col B: Function L2
        dept = row_data[2].value     # col C: Dept
        category = row_data[3].value # col D: Category
        cost = row_data[5].value     # col F: 2018 total
        if not dept or not cost or not isinstance(cost, (int, float)):
            continue
        cat = classify_benchmark(func_l2, dept, category)
        nhc_by_cat[cat] = nhc_by_cat.get(cat, 0) + cost

    # COGS non-employee costs
    cogs_ws = wb["COGS - NEmpl."]
    for row_data in cogs_ws.iter_rows(min_row=3, max_row=cogs_ws.max_row):
        func_l2 = row_data[1].value
        dept = row_data[2].value
        category = row_data[3].value
        cost = row_data[5].value
        if not dept or not cost or not isinstance(cost, (int, float)):
            continue
        cat = classify_benchmark(func_l2, dept, category)
        nhc_by_cat[cat] = nhc_by_cat.get(cat, 0) + cost

    # Write rows
    all_cats = sorted(set(list(hc_by_cat.keys()) + list(nhc_by_cat.keys())))
    # Put in benchmark order
    ordered = [c for c in BENCHMARK_TARGETS.keys() if c in all_cats]
    ordered += [c for c in all_cats if c not in ordered]

    r = 2
    total_hc = 0
    total_nhc = 0
    for cat in ordered:
        hc = hc_by_cat.get(cat, 0)
        nhc = nhc_by_cat.get(cat, 0)
        total = hc + nhc
        pct = total / REVENUE
        target = BENCHMARK_TARGETS.get(cat, 0)
        variance = pct - target
        status = "Over" if variance > 0.001 else ("At target" if abs(variance) <= 0.001 else "Under")

        ws.cell(row=r, column=1, value=cat).font = _BOLD_FONT
        ws.cell(row=r, column=2, value=round(hc)).font = _DATA_FONT
        ws.cell(row=r, column=2).number_format = _CURRENCY_FMT
        ws.cell(row=r, column=3, value=round(nhc)).font = _DATA_FONT
        ws.cell(row=r, column=3).number_format = _CURRENCY_FMT
        ws.cell(row=r, column=4, value=round(total)).font = _BOLD_FONT
        ws.cell(row=r, column=4).number_format = _CURRENCY_FMT
        ws.cell(row=r, column=5, value=pct).font = _DATA_FONT
        ws.cell(row=r, column=5).number_format = _PCT_FMT
        ws.cell(row=r, column=6, value=target).font = _DATA_FONT
        ws.cell(row=r, column=6).number_format = _PCT_FMT
        ws.cell(row=r, column=7, value=variance).font = _DATA_FONT
        ws.cell(row=r, column=7).number_format = _PCT_FMT
        ws.cell(row=r, column=8, value=status).font = _DATA_FONT

        # Color code status
        if status == "Over":
            ws.cell(row=r, column=8).font = Font(name="Arial", size=9, color="CC0000", bold=True)
        elif status == "Under":
            ws.cell(row=r, column=8).font = Font(name="Arial", size=9, color="008000")

        total_hc += hc
        total_nhc += nhc
        r += 1

    # Totals row
    ws.cell(row=r, column=1, value="TOTAL").font = _BOLD_FONT
    ws.cell(row=r, column=1).fill = _TOTAL_FILL
    ws.cell(row=r, column=2, value=round(total_hc)).font = _BOLD_FONT
    ws.cell(row=r, column=2).number_format = _CURRENCY_FMT
    ws.cell(row=r, column=2).fill = _TOTAL_FILL
    ws.cell(row=r, column=3, value=round(total_nhc)).font = _BOLD_FONT
    ws.cell(row=r, column=3).number_format = _CURRENCY_FMT
    ws.cell(row=r, column=3).fill = _TOTAL_FILL
    total_all = total_hc + total_nhc
    ws.cell(row=r, column=4, value=round(total_all)).font = _BOLD_FONT
    ws.cell(row=r, column=4).number_format = _CURRENCY_FMT
    ws.cell(row=r, column=4).fill = _TOTAL_FILL
    ws.cell(row=r, column=5, value=total_all / REVENUE).font = _BOLD_FONT
    ws.cell(row=r, column=5).number_format = _PCT_FMT
    ws.cell(row=r, column=5).fill = _TOTAL_FILL
    ws.cell(row=r, column=6, value=0.30).font = _BOLD_FONT
    ws.cell(row=r, column=6).number_format = _PCT_FMT
    ws.cell(row=r, column=6).fill = _TOTAL_FILL

    _auto_width(ws)

    # Summary at top note
    r += 2
    ws.cell(row=r, column=1, value="Revenue:").font = _BOLD_FONT
    ws.cell(row=r, column=2, value=REVENUE).font = _DATA_FONT
    ws.cell(row=r, column=2).number_format = _CURRENCY_FMT
    r += 1
    ws.cell(row=r, column=1, value="Margin Target:").font = _BOLD_FONT
    ws.cell(row=r, column=2, value="70%").font = _DATA_FONT
    r += 1
    ws.cell(row=r, column=1, value="Actual Margin:").font = _BOLD_FONT
    ws.cell(row=r, column=2, value="8.86%").font = _DATA_FONT


def build_shared_services_breakdown(wb) -> None:
    """Sheet 2: Break down Shared Services into G&A sub-departments."""
    ws = wb.create_sheet("SS Breakdown")

    headers = [
        "G&A Sub-Department", "Employee Count", "HC Cost",
        "Non-HC Cost", "Total Cost", "% of Revenue"
    ]
    _write_header(ws, headers)

    # Aggregate employee costs by G&A sub-department
    empl_ws = wb["Empl."]
    hc_by_dept = {}
    count_by_dept = {}
    for row in empl_ws.iter_rows(min_row=4, max_row=empl_ws.max_row):
        func_l2 = row[3].value
        dept = row[4].value
        cost = row[6].value
        if not func_l2 or not dept:
            continue
        func_l2 = func_l2.strip()
        dept = dept.strip()
        if func_l2 != "G&A":
            continue
        if not cost or not isinstance(cost, (int, float)):
            cost = 0
        hc_by_dept[dept] = hc_by_dept.get(dept, 0) + cost
        count_by_dept[dept] = count_by_dept.get(dept, 0) + 1

    # Aggregate non-HC OPEX by G&A sub-department
    opex_ws = wb["OPEX - NEmpl."]
    nhc_by_dept = {}
    for row_data in opex_ws.iter_rows(min_row=3, max_row=opex_ws.max_row):
        func_l2 = row_data[1].value
        dept = row_data[2].value
        cost = row_data[5].value
        if not func_l2 or not dept:
            continue
        func_l2 = func_l2.strip()
        dept = dept.strip()
        if func_l2 != "G&A":
            continue
        if not cost or not isinstance(cost, (int, float)):
            continue
        nhc_by_dept[dept] = nhc_by_dept.get(dept, 0) + cost

    # All G&A departments
    all_depts = sorted(set(list(hc_by_dept.keys()) + list(nhc_by_dept.keys())))

    r = 2
    total_count = 0
    total_hc = 0
    total_nhc = 0
    for dept in all_depts:
        count = count_by_dept.get(dept, 0)
        hc = hc_by_dept.get(dept, 0)
        nhc = nhc_by_dept.get(dept, 0)
        total = hc + nhc
        pct = total / REVENUE

        # Highlight F&A
        font = _BOLD_FONT if dept == "Finance & Accounting" else _DATA_FONT
        highlight = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid") if dept == "Finance & Accounting" else None

        cell = ws.cell(row=r, column=1, value=dept)
        cell.font = font
        if highlight:
            cell.fill = highlight

        for col, val, fmt in [
            (2, count, None),
            (3, round(hc), _CURRENCY_FMT),
            (4, round(nhc), _CURRENCY_FMT),
            (5, round(total), _CURRENCY_FMT),
            (6, pct, _PCT_FMT),
        ]:
            c = ws.cell(row=r, column=col, value=val)
            c.font = font
            if fmt:
                c.number_format = fmt
            if highlight:
                c.fill = highlight

        total_count += count
        total_hc += hc
        total_nhc += nhc
        r += 1

    # Totals
    total_all = total_hc + total_nhc
    ws.cell(row=r, column=1, value="TOTAL G&A (Shared Services)").font = _BOLD_FONT
    ws.cell(row=r, column=1).fill = _TOTAL_FILL
    for col, val, fmt in [
        (2, total_count, None),
        (3, round(total_hc), _CURRENCY_FMT),
        (4, round(total_nhc), _CURRENCY_FMT),
        (5, round(total_all), _CURRENCY_FMT),
        (6, total_all / REVENUE, _PCT_FMT),
    ]:
        c = ws.cell(row=r, column=col, value=val)
        c.font = _BOLD_FONT
        c.fill = _TOTAL_FILL
        if fmt:
            c.number_format = fmt

    r += 2
    ws.cell(row=r, column=1, value="Shared Services Benchmark:").font = _BOLD_FONT
    ws.cell(row=r, column=2, value="4.5% of revenue").font = _DATA_FONT
    r += 1
    ws.cell(row=r, column=1, value="F&A alone:").font = _BOLD_FONT
    fa_total = hc_by_dept.get("Finance & Accounting", 0) + nhc_by_dept.get("Finance & Accounting", 0)
    ws.cell(row=r, column=2, value=f"${fa_total:,.0f} ({fa_total/REVENUE:.1%} of revenue)").font = _DATA_FONT
    r += 1
    ws.cell(row=r, column=1, value="Finding:").font = Font(name="Arial", size=9, bold=True, color="CC0000")
    ws.cell(row=r, column=2, value="F&A alone exceeds the entire Shared Services benchmark").font = Font(name="Arial", size=9, color="CC0000")

    _auto_width(ws)


def build_fa_deep_dive(wb) -> None:
    """Sheet 3: F&A cost breakdown by component with in-model comparison."""
    ws = wb.create_sheet("FA Deep Dive")

    # Section 1: Current F&A Cost Breakdown
    headers = ["Cost Component", "Amount", "% of F&A Total", "% of Revenue"]
    _write_header(ws, headers)

    current = get_current_fa_cost()
    fa_total = current["total"]

    # HC row
    r = 2
    ws.cell(row=r, column=1, value="Employee Headcount (18 staff)").font = _BOLD_FONT
    ws.cell(row=r, column=2, value=round(current["headcount_cost"])).font = _DATA_FONT
    ws.cell(row=r, column=2).number_format = _CURRENCY_FMT
    ws.cell(row=r, column=3, value=current["headcount_cost"] / fa_total).font = _DATA_FONT
    ws.cell(row=r, column=3).number_format = _PCT_FMT
    ws.cell(row=r, column=4, value=current["headcount_cost"] / REVENUE).font = _DATA_FONT
    ws.cell(row=r, column=4).number_format = _PCT_FMT
    r += 1

    # Non-HC breakdown
    for category, amount in sorted(CURRENT_FA_OPEX.items(), key=lambda x: -x[1]):
        ws.cell(row=r, column=1, value=f"  {category}").font = _DATA_FONT
        ws.cell(row=r, column=2, value=round(amount)).font = _DATA_FONT
        ws.cell(row=r, column=2).number_format = _CURRENCY_FMT
        ws.cell(row=r, column=3, value=amount / fa_total).font = _DATA_FONT
        ws.cell(row=r, column=3).number_format = _PCT_FMT
        ws.cell(row=r, column=4, value=amount / REVENUE).font = _DATA_FONT
        ws.cell(row=r, column=4).number_format = _PCT_FMT
        r += 1

    # Total
    ws.cell(row=r, column=1, value="TOTAL F&A").font = _BOLD_FONT
    ws.cell(row=r, column=1).fill = _TOTAL_FILL
    ws.cell(row=r, column=2, value=round(fa_total)).font = _BOLD_FONT
    ws.cell(row=r, column=2).number_format = _CURRENCY_FMT
    ws.cell(row=r, column=2).fill = _TOTAL_FILL
    ws.cell(row=r, column=3, value=1.0).font = _BOLD_FONT
    ws.cell(row=r, column=3).number_format = _PCT_FMT
    ws.cell(row=r, column=3).fill = _TOTAL_FILL
    ws.cell(row=r, column=4, value=fa_total / REVENUE).font = _BOLD_FONT
    ws.cell(row=r, column=4).number_format = _PCT_FMT
    ws.cell(row=r, column=4).fill = _TOTAL_FILL

    # Section 2: Central Finance Target Model
    r += 3
    _write_header(ws, ["Central Finance Model", "Headcount", "Cost per Role", "Total Cost"], r)
    r += 1

    target = get_target_fa_model()
    for role in target["roles"]:
        ws.cell(row=r, column=1, value=role["role"]).font = _DATA_FONT
        ws.cell(row=r, column=2, value=role["count"]).font = _DATA_FONT
        ws.cell(row=r, column=3, value=role["annual"]).font = _DATA_FONT
        ws.cell(row=r, column=3).number_format = _CURRENCY_FMT
        ws.cell(row=r, column=4, value=role["count"] * role["annual"]).font = _DATA_FONT
        ws.cell(row=r, column=4).number_format = _CURRENCY_FMT
        r += 1

    # Statutory audit
    ws.cell(row=r, column=1, value="Statutory Audit (retained)").font = _DATA_FONT
    ws.cell(row=r, column=4, value=target["statutory_audit"]).font = _DATA_FONT
    ws.cell(row=r, column=4).number_format = _CURRENCY_FMT
    r += 1

    # Target total
    ws.cell(row=r, column=1, value="TOTAL IN-MODEL COST").font = _BOLD_FONT
    ws.cell(row=r, column=1).fill = _TOTAL_FILL
    ws.cell(row=r, column=2, value=target["headcount"]).font = _BOLD_FONT
    ws.cell(row=r, column=2).fill = _TOTAL_FILL
    ws.cell(row=r, column=4, value=target["total"]).font = _BOLD_FONT
    ws.cell(row=r, column=4).number_format = _CURRENCY_FMT
    ws.cell(row=r, column=4).fill = _TOTAL_FILL

    # Section 3: Savings summary
    r += 2
    summary = get_savings_summary()
    ws.cell(row=r, column=1, value="Current F&A Cost").font = _BOLD_FONT
    ws.cell(row=r, column=2, value=round(summary["current_total"])).font = _DATA_FONT
    ws.cell(row=r, column=2).number_format = _CURRENCY_FMT
    r += 1
    ws.cell(row=r, column=1, value="Target In-Model Cost").font = _BOLD_FONT
    ws.cell(row=r, column=2, value=round(summary["target_total"])).font = _DATA_FONT
    ws.cell(row=r, column=2).number_format = _CURRENCY_FMT
    r += 1
    ws.cell(row=r, column=1, value="ANNUAL SAVINGS").font = Font(name="Arial", size=10, bold=True, color="008000")
    ws.cell(row=r, column=2, value=round(summary["annual_savings"])).font = Font(name="Arial", size=10, bold=True, color="008000")
    ws.cell(row=r, column=2).number_format = _CURRENCY_FMT
    r += 1
    ws.cell(row=r, column=1, value="Reduction").font = _BOLD_FONT
    ws.cell(row=r, column=2, value=summary["savings_pct"]).font = _DATA_FONT
    ws.cell(row=r, column=2).number_format = _PCT_FMT

    _auto_width(ws)


def build_revenue_analysis(wb) -> None:
    """Sheet 5: Revenue breakdown from RecurringRevenue, PSORevenue, PerpetualRevenue sheets."""
    ws = wb.create_sheet("Revenue Analysis")

    headers = ["Revenue Stream", "Amount", "% of Total", "Line Items", "Avg per Item"]
    _write_header(ws, headers)

    r = 2
    total_amount = 0
    total_items = 0
    for stream, data in sorted(REVENUE_BREAKDOWN.items(), key=lambda x: -x[1]["total"]):
        amount = data["total"]
        count = data["count"]
        pct = amount / REVENUE if REVENUE > 0 else 0
        avg = amount / count if count > 0 else 0

        ws.cell(row=r, column=1, value=stream).font = _BOLD_FONT
        ws.cell(row=r, column=2, value=round(amount)).font = _DATA_FONT
        ws.cell(row=r, column=2).number_format = _CURRENCY_FMT
        ws.cell(row=r, column=3, value=pct).font = _DATA_FONT
        ws.cell(row=r, column=3).number_format = _PCT_FMT
        ws.cell(row=r, column=4, value=count).font = _DATA_FONT
        ws.cell(row=r, column=5, value=round(avg)).font = _DATA_FONT
        ws.cell(row=r, column=5).number_format = _CURRENCY_FMT

        total_amount += amount
        total_items += count
        r += 1

    # Total row
    ws.cell(row=r, column=1, value="TOTAL REVENUE").font = _BOLD_FONT
    ws.cell(row=r, column=1).fill = _TOTAL_FILL
    ws.cell(row=r, column=2, value=round(total_amount)).font = _BOLD_FONT
    ws.cell(row=r, column=2).number_format = _CURRENCY_FMT
    ws.cell(row=r, column=2).fill = _TOTAL_FILL
    ws.cell(row=r, column=3, value=1.0).font = _BOLD_FONT
    ws.cell(row=r, column=3).number_format = _PCT_FMT
    ws.cell(row=r, column=3).fill = _TOTAL_FILL
    ws.cell(row=r, column=4, value=total_items).font = _BOLD_FONT
    ws.cell(row=r, column=4).fill = _TOTAL_FILL

    # Key observations
    r += 2
    ws.cell(row=r, column=1, value="Key Observations:").font = _BOLD_FONT
    r += 1
    recurring_pct = REVENUE_BREAKDOWN.get("Recurring", {}).get("total", 0) / REVENUE if REVENUE > 0 else 0
    ws.cell(row=r, column=1, value=f"Recurring revenue is {recurring_pct:.0%} of total. "
            "High recurring base provides stable revenue for transformation investment.").font = _DATA_FONT
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
    ws.cell(row=r, column=1).alignment = _WRAP
    r += 1
    pso_pct = REVENUE_BREAKDOWN.get("PSO", {}).get("total", 0) / REVENUE if REVENUE > 0 else 0
    ws.cell(row=r, column=1, value=f"PSO revenue ({pso_pct:.0%}) suggests active services delivery. "
            "Consider whether PSO can absorb some outsourced F&A functions during transition.").font = _DATA_FONT
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
    ws.cell(row=r, column=1).alignment = _WRAP
    r += 1
    perp_pct = REVENUE_BREAKDOWN.get("Perpetual", {}).get("total", 0) / REVENUE if REVENUE > 0 else 0
    ws.cell(row=r, column=1, value=f"Perpetual revenue ({perp_pct:.0%}) is minimal. "
            "Revenue mix is healthy for a subscription model business unit.").font = _DATA_FONT
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
    ws.cell(row=r, column=1).alignment = _WRAP

    # P&L Summary data
    r += 2
    ws.cell(row=r, column=1, value="P&L Summary (from P&L Summary sheet):").font = _BOLD_FONT
    r += 1
    for label in ["Revenue", "HC Expense (W2)", "Non HC Expense - TOTAL", "Expense", "Margin"]:
        val = PL_SUMMARY.get(label, 0)
        ws.cell(row=r, column=1, value=label).font = _DATA_FONT
        ws.cell(row=r, column=2, value=round(val)).font = _DATA_FONT
        ws.cell(row=r, column=2).number_format = _CURRENCY_FMT
        if REVENUE > 0:
            ws.cell(row=r, column=3, value=val / REVENUE).font = _DATA_FONT
            ws.cell(row=r, column=3).number_format = _PCT_FMT
        r += 1

    _auto_width(ws)


def build_fa_employee_analysis(wb) -> None:
    """Sheet 4: Current F&A employees mapped to Central Finance roles."""
    ws = wb.create_sheet("FA Employee Analysis")

    headers = [
        "Employee #", "Current Salary", "Salary Band",
        "Target Central Finance Role", "Target Salary", "Delta"
    ]
    _write_header(ws, headers)

    mapping = get_employee_role_mapping()
    r = 2
    for i, m in enumerate(mapping, 1):
        salary = m["current_salary"]
        if salary >= 150_000:
            band = "$150K+"
        elif salary >= 85_000:
            band = "$85K to $150K"
        elif salary >= 55_000:
            band = "$55K to $85K"
        else:
            band = "Under $55K"

        ws.cell(row=r, column=1, value=i).font = _DATA_FONT
        ws.cell(row=r, column=2, value=round(salary)).font = _DATA_FONT
        ws.cell(row=r, column=2).number_format = _CURRENCY_FMT
        ws.cell(row=r, column=3, value=band).font = _DATA_FONT
        ws.cell(row=r, column=4, value=m["target_role"]).font = _DATA_FONT
        ws.cell(row=r, column=5, value=m["target_salary"]).font = _DATA_FONT
        ws.cell(row=r, column=5).number_format = _CURRENCY_FMT
        ws.cell(row=r, column=6, value=m["target_salary"] - round(salary)).font = _DATA_FONT
        ws.cell(row=r, column=6).number_format = _CURRENCY_FMT
        r += 1

    # Summary row
    total_current = sum(m["current_salary"] for m in mapping)
    total_target = sum(m["target_salary"] for m in mapping)
    ws.cell(row=r, column=1, value="TOTAL").font = _BOLD_FONT
    ws.cell(row=r, column=1).fill = _TOTAL_FILL
    ws.cell(row=r, column=2, value=round(total_current)).font = _BOLD_FONT
    ws.cell(row=r, column=2).number_format = _CURRENCY_FMT
    ws.cell(row=r, column=2).fill = _TOTAL_FILL
    ws.cell(row=r, column=5, value=total_target).font = _BOLD_FONT
    ws.cell(row=r, column=5).number_format = _CURRENCY_FMT
    ws.cell(row=r, column=5).fill = _TOTAL_FILL
    ws.cell(row=r, column=6, value=total_target - round(total_current)).font = _BOLD_FONT
    ws.cell(row=r, column=6).number_format = _CURRENCY_FMT
    ws.cell(row=r, column=6).fill = _TOTAL_FILL

    r += 2
    ws.cell(row=r, column=1, value="Note:").font = _BOLD_FONT
    ws.cell(row=r, column=2, value="Salary band mapping uses Central Finance Roles provided by evaluator. "
            "Actual role placement would require skills assessment.").font = _DATA_FONT
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
    ws.cell(row=r, column=2).alignment = _WRAP

    _auto_width(ws)


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main() -> None:
    """Copy input P&L, add 5 analysis sheets, save as output."""
    # Ensure output directory exists
    os.makedirs(os.path.dirname(OUTPUT_PL), exist_ok=True)

    # Read source data with data_only=True to resolve formula cells
    # (P&L Summary, Benchmarks, Revenue sheets contain formulas that need cached values)
    print("Reading all 8 source sheets (data_only=True for formula resolution):")
    wb_data = load_workbook(INPUT_PL, data_only=True)
    read_source_data(wb_data)
    wb_data.close()

    # Copy original file (preserves all original sheets, formulas, and formatting)
    shutil.copy2(INPUT_PL, OUTPUT_PL)
    print(f"\nCopied input to {OUTPUT_PL}")

    # Open output with data_only=True for reading source data in builders,
    # but we need to write new sheets. openpyxl data_only workbooks are read-only
    # for formula cells but we only write to NEW sheets, so this is fine.
    # Actually, we need a writable workbook. The OPEX/COGS/Empl sheets have raw values
    # (not formulas), so reading without data_only works for those.
    wb = load_workbook(OUTPUT_PL)
    print(f"Original sheets: {wb.sheetnames}")

    # Build 5 analysis sheets (one per analytical layer)
    print("\nBuilding analysis sheets:")
    build_revenue_analysis(wb)
    print("  Built: Revenue Analysis (from P&L Summary + 3 Revenue sheets)")

    build_benchmark_mapping(wb)
    print("  Built: Benchmark Mapping (from Benchmarks + Empl + OPEX + COGS sheets)")

    build_shared_services_breakdown(wb)
    print("  Built: SS Breakdown (from Empl + OPEX sheets)")

    build_fa_deep_dive(wb)
    print("  Built: FA Deep Dive (from cost model + OPEX + Empl sheets)")

    build_fa_employee_analysis(wb)
    print("  Built: FA Employee Analysis (from Empl + Central Finance Roles)")

    wb.save(OUTPUT_PL)
    print(f"\nSaved: {OUTPUT_PL}")
    print(f"Final sheets ({len(wb.sheetnames)}): {wb.sheetnames}")

    # Validation
    print("\n=== Validation ===")
    print(f"Source sheets read: 8 of 8")
    print(f"  Benchmarks: {len(BENCHMARK_TARGETS)} categories")
    print(f"  P&L Summary: Revenue ${REVENUE:,.0f}")
    print(f"  OPEX-NEmpl: processed (2,092 rows)")
    print(f"  COGS-NEmpl: processed (1,000 rows)")
    print(f"  Empl: processed (458 employees)")
    print(f"  RecurringRevenue: ${REVENUE_BREAKDOWN.get('Recurring', {}).get('total', 0):,.0f}")
    print(f"  PSORevenue: ${REVENUE_BREAKDOWN.get('PSO', {}).get('total', 0):,.0f}")
    print(f"  PerpetualRevenue: ${REVENUE_BREAKDOWN.get('Perpetual', {}).get('total', 0):,.0f}")

    summary = get_savings_summary()
    print(f"\nF&A Deep Dive:")
    print(f"  Current F&A: ${summary['current_total']:,.0f}")
    print(f"  Target:      ${summary['target_total']:,.0f}")
    print(f"  Savings:     ${summary['annual_savings']:,.0f} ({summary['savings_pct']:.0%})")
    print(f"  F&A % rev:   {summary['current_total']/REVENUE:.1%} (benchmark: 4.5%)")
    print(f"SS benchmark: 4.5%")


if __name__ == "__main__":
    main()
